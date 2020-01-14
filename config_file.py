import datetime
import pandas as pd
import openpyxl

def read_config():
    OAuth_codes = pd.read_excel('./slackbot_config.xlsx', 'OAuth_codes')
    messaging_queue = pd.read_excel('./slackbot_config.xlsx', 'messaging_queue')
    OAuth_codes = dict(zip(OAuth_codes.Client,OAuth_codes.OAuth_Access_Token ))
    return (OAuth_codes, messaging_queue)

def update_messaging_status():
    messaging_queue = pd.read_excel('./slackbot_config.xlsx', 'messaging_queue')
    col_idx = [col for col in messaging_queue.columns].index('sent')
    length = messaging_queue.message.count()
    wb= openpyxl.load_workbook('slackbot_config.xlsx')
    messaging_queue_ws = wb.worksheets[1]

    for i in range(length):
        messaging_queue_ws.cell(i+2 ,column = col_idx+1 ).value = True
        messaging_queue_ws.cell(i+2 ,column = col_idx+2 ).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb.save('./slackbot_config.xlsx')